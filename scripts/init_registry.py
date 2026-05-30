"""Initialize a blank or mock registry Excel workbook.

Usage:
    python -m scripts.init_registry --output registry/mock_data.xlsx
    python -m scripts.init_registry --output registry/my_registry.xlsx --with-mock
"""

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

SHEET_HEADERS = {
    "Domain": ["code", "name", "parent_code", "description", "source"],
    "Table": ["fqn", "schema_name", "table_name", "type", "business_object", "domains", "comment", "status"],
    "Column": [
        "fqn", "table_fqn", "name", "data_type", "nullable",
        "is_pk", "is_fk", "ref_column_fqn", "semantic_type", "domains", "comment",
    ],
    "Relationship": [
        "src_fqn", "dst_fqn", "node_level", "rel_type",
        "is_directed", "properties", "source", "status",
    ],
}


def _write_header(ws, headers: list[str]):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"


def create_blank_workbook(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Domain"
    _write_header(ws, SHEET_HEADERS["Domain"])

    for sheet_name in ["Table", "Column", "Relationship"]:
        ws = wb.create_sheet(sheet_name)
        _write_header(ws, SHEET_HEADERS[sheet_name])

    wb.save(path)
    return path


MOCK_DOMAINS = [
    ["procurement", "采购域", "", "采购核心业务流程", "manual"],
    ["supply_base", "供应商管理", "procurement", "供应商全生命周期管理", "manual"],
    ["po_management", "采购订单管理", "procurement", "采购订单与合同管理", "manual"],
    ["finance", "财务域", "", "财务结算与核销", "manual"],
]

MOCK_TABLES = [
    ["proc.public.supplier", "public", "supplier", "table", "供应商主数据", "supply_base", "供应商基本信息表", "active"],
    ["proc.public.supplier_qual", "public", "supplier_qualification", "table", "供应商资质", "supply_base", "供应商资质文件表", "active"],
    ["proc.public.po_order", "public", "po_order", "table", "采购订单", "po_management", "采购订单头表", "active"],
    ["proc.public.po_line", "public", "po_line", "table", "采购订单行", "po_management", "采购订单行项目表", "active"],
    ["proc.public.contract", "public", "contract", "table", "采购合同", "po_management", "采购合同主表", "active"],
    ["proc.public.goods_receipt", "public", "goods_receipt", "table", "收货单", "po_management", "入库收货单", "active"],
    ["proc.public.invoice", "public", "invoice", "table", "发票", "finance", "采购发票表", "active"],
    ["proc.public.payment", "public", "payment", "table", "付款记录", "finance", "付款核销记录", "active"],
]

MOCK_COLUMNS = [
    # supplier
    ["proc.public.supplier.id", "proc.public.supplier", "id", "bigint", "false", "true", "false", "", "ID", "supply_base", "主键"],
    ["proc.public.supplier.code", "proc.public.supplier", "code", "varchar(32)", "false", "false", "false", "", "供应商编码", "supply_base", "供应商编码"],
    ["proc.public.supplier.name", "proc.public.supplier", "name", "varchar(200)", "false", "false", "false", "", "名称", "supply_base", "供应商名称"],
    ["proc.public.supplier.status", "proc.public.supplier", "status", "varchar(16)", "false", "false", "false", "", "状态码", "supply_base", "供应商状态"],
    ["proc.public.supplier.created_at", "proc.public.supplier", "created_at", "timestamp", "false", "false", "false", "", "时间", "supply_base", "创建时间"],
    # supplier_qual
    ["proc.public.supplier_qual.id", "proc.public.supplier_qual", "id", "bigint", "false", "true", "false", "", "ID", "supply_base", "主键"],
    ["proc.public.supplier_qual.supplier_id", "proc.public.supplier_qual", "supplier_id", "bigint", "false", "false", "true", "proc.public.supplier.id", "外键", "supply_base", "供应商ID"],
    ["proc.public.supplier_qual.qual_type", "proc.public.supplier_qual", "qual_type", "varchar(64)", "false", "false", "false", "", "分类", "supply_base", "资质类型"],
    ["proc.public.supplier_qual.expire_date", "proc.public.supplier_qual", "expire_date", "date", "true", "false", "false", "", "时间", "supply_base", "有效期至"],
    # po_order
    ["proc.public.po_order.id", "proc.public.po_order", "id", "bigint", "false", "true", "false", "", "ID", "po_management", "主键"],
    ["proc.public.po_order.order_no", "proc.public.po_order", "order_no", "varchar(32)", "false", "false", "false", "", "单据编码", "po_management", "采购订单编号"],
    ["proc.public.po_order.supplier_id", "proc.public.po_order", "supplier_id", "bigint", "false", "false", "true", "proc.public.supplier.id", "外键", "po_management,supply_base", "供应商ID"],
    ["proc.public.po_order.contract_id", "proc.public.po_order", "contract_id", "bigint", "true", "false", "true", "proc.public.contract.id", "外键", "po_management", "合同ID"],
    ["proc.public.po_order.total_amt", "proc.public.po_order", "total_amt", "decimal(18,2)", "false", "false", "false", "", "金额", "po_management", "订单总金额"],
    ["proc.public.po_order.currency", "proc.public.po_order", "currency", "varchar(8)", "false", "false", "false", "", "枚举", "po_management", "币种"],
    ["proc.public.po_order.status", "proc.public.po_order", "status", "varchar(16)", "false", "false", "false", "", "状态码", "po_management", "订单状态"],
    ["proc.public.po_order.created_at", "proc.public.po_order", "created_at", "timestamp", "false", "false", "false", "", "时间", "po_management", "创建时间"],
    # po_line
    ["proc.public.po_line.id", "proc.public.po_line", "id", "bigint", "false", "true", "false", "", "ID", "po_management", "主键"],
    ["proc.public.po_line.po_order_id", "proc.public.po_line", "po_order_id", "bigint", "false", "false", "true", "proc.public.po_order.id", "外键", "po_management", "所属采购订单ID"],
    ["proc.public.po_line.line_no", "proc.public.po_line", "line_no", "int", "false", "false", "false", "", "序号", "po_management", "行号"],
    ["proc.public.po_line.material_code", "proc.public.po_line", "material_code", "varchar(64)", "false", "false", "false", "", "物料编码", "po_management", "物料编码"],
    ["proc.public.po_line.quantity", "proc.public.po_line", "quantity", "decimal(18,4)", "false", "false", "false", "", "数量", "po_management", "采购数量"],
    ["proc.public.po_line.unit_price", "proc.public.po_line", "unit_price", "decimal(18,4)", "false", "false", "false", "", "金额", "po_management", "单价"],
    ["proc.public.po_line.line_amt", "proc.public.po_line", "line_amt", "decimal(18,2)", "false", "false", "false", "", "金额", "po_management", "行金额"],
    # contract
    ["proc.public.contract.id", "proc.public.contract", "id", "bigint", "false", "true", "false", "", "ID", "po_management", "主键"],
    ["proc.public.contract.contract_no", "proc.public.contract", "contract_no", "varchar(32)", "false", "false", "false", "", "单据编码", "po_management", "合同编号"],
    ["proc.public.contract.supplier_id", "proc.public.contract", "supplier_id", "bigint", "false", "false", "true", "proc.public.supplier.id", "外键", "po_management,supply_base", "供应商ID"],
    ["proc.public.contract.total_amt", "proc.public.contract", "total_amt", "decimal(18,2)", "false", "false", "false", "", "金额", "po_management", "合同金额"],
    ["proc.public.contract.start_date", "proc.public.contract", "start_date", "date", "false", "false", "false", "", "时间", "po_management", "生效日期"],
    ["proc.public.contract.end_date", "proc.public.contract", "end_date", "date", "false", "false", "false", "", "时间", "po_management", "到期日期"],
    # goods_receipt
    ["proc.public.goods_receipt.id", "proc.public.goods_receipt", "id", "bigint", "false", "true", "false", "", "ID", "po_management", "主键"],
    ["proc.public.goods_receipt.po_line_id", "proc.public.goods_receipt", "po_line_id", "bigint", "false", "false", "true", "proc.public.po_line.id", "外键", "po_management", "采购订单行ID"],
    ["proc.public.goods_receipt.receipt_qty", "proc.public.goods_receipt", "receipt_qty", "decimal(18,4)", "false", "false", "false", "", "数量", "po_management", "实收数量"],
    ["proc.public.goods_receipt.receipt_date", "proc.public.goods_receipt", "receipt_date", "date", "false", "false", "false", "", "时间", "po_management", "收货日期"],
    # invoice
    ["proc.public.invoice.id", "proc.public.invoice", "id", "bigint", "false", "true", "false", "", "ID", "finance", "主键"],
    ["proc.public.invoice.invoice_no", "proc.public.invoice", "invoice_no", "varchar(32)", "false", "false", "false", "", "单据编码", "finance", "发票号码"],
    ["proc.public.invoice.po_order_id", "proc.public.invoice", "po_order_id", "bigint", "true", "false", "true", "proc.public.po_order.id", "外键", "finance,po_management", "采购订单ID"],
    ["proc.public.invoice.amount", "proc.public.invoice", "amount", "decimal(18,2)", "false", "false", "false", "", "金额", "finance", "发票金额"],
    ["proc.public.invoice.tax_amt", "proc.public.invoice", "tax_amt", "decimal(18,2)", "false", "false", "false", "", "金额", "finance", "税额"],
    ["proc.public.invoice.invoice_date", "proc.public.invoice", "invoice_date", "date", "false", "false", "false", "", "时间", "finance", "开票日期"],
    # payment
    ["proc.public.payment.id", "proc.public.payment", "id", "bigint", "false", "true", "false", "", "ID", "finance", "主键"],
    ["proc.public.payment.invoice_id", "proc.public.payment", "invoice_id", "bigint", "false", "false", "true", "proc.public.invoice.id", "外键", "finance", "发票ID"],
    ["proc.public.payment.pay_amt", "proc.public.payment", "pay_amt", "decimal(18,2)", "false", "false", "false", "", "金额", "finance", "付款金额"],
    ["proc.public.payment.pay_date", "proc.public.payment", "pay_date", "date", "false", "false", "false", "", "时间", "finance", "付款日期"],
]

MOCK_RELATIONSHIPS = [
    # Physical foreign keys
    ["proc.public.supplier_qual.supplier_id", "proc.public.supplier.id", "column", "REFERENCES", "true", "", "introspect", "active"],
    ["proc.public.po_order.supplier_id", "proc.public.supplier.id", "column", "REFERENCES", "true", "", "introspect", "active"],
    ["proc.public.po_order.contract_id", "proc.public.contract.id", "column", "REFERENCES", "true", "", "introspect", "active"],
    ["proc.public.po_line.po_order_id", "proc.public.po_order.id", "column", "REFERENCES", "true", "", "introspect", "active"],
    ["proc.public.goods_receipt.po_line_id", "proc.public.po_line.id", "column", "REFERENCES", "true", "", "introspect", "active"],
    ["proc.public.invoice.po_order_id", "proc.public.po_order.id", "column", "REFERENCES", "true", "", "introspect", "active"],
    ["proc.public.payment.invoice_id", "proc.public.invoice.id", "column", "REFERENCES", "true", "", "introspect", "active"],
    ["proc.public.contract.supplier_id", "proc.public.supplier.id", "column", "REFERENCES", "true", "", "introspect", "active"],
]


def write_mock_data(path: Path) -> Path:
    from openpyxl import load_workbook

    wb = load_workbook(path)

    ws = wb["Domain"]
    for row in MOCK_DOMAINS:
        ws.append(row)

    ws = wb["Table"]
    for row in MOCK_TABLES:
        ws.append(row)

    ws = wb["Column"]
    for row in MOCK_COLUMNS:
        ws.append(row)

    ws = wb["Relationship"]
    for row in MOCK_RELATIONSHIPS:
        ws.append(row)

    wb.save(path)
    return path


def main():
    parser = argparse.ArgumentParser(description="Initialize registry Excel workbook")
    parser.add_argument("--output", "-o", default="registry/mock_data.xlsx", help="Output path")
    parser.add_argument("--with-mock", action="store_true", help="Write mock procurement data")
    args = parser.parse_args()

    path = Path(args.output)
    path.parent.mkdir(parents=True, exist_ok=True)

    create_blank_workbook(path)
    print(f"Blank registry created: {path}")

    if args.with_mock:
        write_mock_data(path)
        print(f"Mock data written ({len(MOCK_DOMAINS)} domains, {len(MOCK_TABLES)} tables, "
              f"{len(MOCK_COLUMNS)} columns, {len(MOCK_RELATIONSHIPS)} relationships)")


if __name__ == "__main__":
    main()
