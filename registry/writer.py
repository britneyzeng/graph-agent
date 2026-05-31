from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from registry.models import (
    DomainDef,
    EntityDef,
    PropertyDef,
    RegistryData,
    RelationshipDef,
)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


class RegistryWriter:
    def __init__(self, xlsx_path: str | Path):
        self.xlsx_path = Path(xlsx_path)

    def write(self, data: RegistryData) -> Path:
        wb = Workbook()

        self._write_domains(wb, data.domains)
        self._write_entities(wb, data.entities)
        self._write_properties(wb, data.properties)
        self._write_relationships(wb, data.relationships)

        wb.save(self.xlsx_path)
        return self.xlsx_path

    def _write_header(self, ws, headers: list[str]):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        ws.freeze_panes = "A2"

    @staticmethod
    def _to_json(val) -> str:
        return json.dumps(val, ensure_ascii=False) if val else ""

    @staticmethod
    def _to_bool(val: bool) -> str:
        return "true" if val else "false"

    def _write_domains(self, wb: Workbook, domains: list[DomainDef]):
        ws = wb.active
        ws.title = "Domain"
        self._write_header(ws, ["code", "name_cn", "name_en", "parent_code", "description", "source", "status"])
        for d in domains:
            ws.append(
                [d.code, d.name_cn, d.name_en or "", d.parent_code or "", d.description, d.source, d.status]
            )

    def _write_entities(self, wb: Workbook, entities: list[EntityDef]):
        ws = wb.create_sheet("Entity")
        self._write_header(
            ws, ["fqn", "entity_type", "name_cn", "name_en", "src_tables", "domains", "description", "source", "status"]
        )
        for e in entities:
            ws.append(
                [
                    e.fqn,
                    e.entity_type,
                    e.name_cn,
                    e.name_en,
                    self._to_json(e.src_tables),
                    self._to_json(e.domains),
                    e.description,
                    e.source,
                    e.status,
                ]
            )

    def _write_properties(self, wb: Workbook, properties: list[PropertyDef]):
        ws = wb.create_sheet("Property")
        self._write_header(
            ws,
            [
                "fqn",
                "entity_fqn",
                "data_type",
                "is_pk",
                "is_fk",
                "ref_property_fqn",
                "description",
                "name_cn",
                "name_en",
                "source",
                "status",
            ],
        )
        for p in properties:
            ws.append(
                [
                    p.fqn,
                    p.entity_fqn,
                    p.data_type,
                    self._to_bool(p.is_pk),
                    self._to_bool(p.is_fk),
                    p.ref_property_fqn or "",
                    p.description,
                    p.name_cn,
                    p.name_en,
                    p.source,
                    p.status,
                ]
            )

    def _write_relationships(self, wb: Workbook, relationships: list[RelationshipDef]):
        ws = wb.create_sheet("Relationship")
        self._write_header(
            ws,
            [
                "src_fqn",
                "dst_fqn",
                "rel_type",
                "is_directed",
                "source",
                "status",
            ],
        )
        for r in relationships:
            ws.append(
                [
                    r.src_fqn,
                    r.dst_fqn,
                    r.rel_type,
                    self._to_bool(r.is_directed),
                    r.source,
                    r.status,
                ]
            )

    def append_relationships(self, relationships: list[RelationshipDef]) -> Path:
        from openpyxl import load_workbook

        wb = load_workbook(self.xlsx_path)
        ws = wb["Relationship"]
        for r in relationships:
            ws.append(
                [
                    r.src_fqn,
                    r.dst_fqn,
                    r.rel_type,
                    self._to_bool(r.is_directed),
                    r.source,
                    r.status,
                ]
            )
        wb.save(self.xlsx_path)
        return self.xlsx_path
