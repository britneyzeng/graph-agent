from __future__ import annotations

import json
import logging
from pathlib import Path

from typing import Any

from openpyxl import load_workbook

from registry.models import (
    DomainDef,
    EntityDef,
    LogicDef,
    PropertyDef,
    RegistryData,
    RelationshipDef,
)

logger = logging.getLogger(__name__)

SHEET_DOMAIN = "Domain"
SHEET_ENTITY = "Entity"
SHEET_PROPERTY = "Property"
SHEET_LOGIC = "Logic"
SHEET_RELATIONSHIP = "Relationship"

REQUIRED_SHEETS = {SHEET_DOMAIN, SHEET_ENTITY, SHEET_PROPERTY, SHEET_RELATIONSHIP}


class RegistryLoader:
    def __init__(self, xlsx_path: str | Path):
        self.xlsx_path = Path(xlsx_path)

    def load(self) -> RegistryData:
        if not self.xlsx_path.exists():
            raise FileNotFoundError(f"Registry file not found: {self.xlsx_path}")

        wb = load_workbook(self.xlsx_path, read_only=True, data_only=True)
        sheet_names = set(wb.sheetnames)
        missing = REQUIRED_SHEETS - sheet_names
        if missing:
            wb.close()
            raise ValueError(f"Missing required sheets: {missing}")

        data = RegistryData(
            domains=self._load_domains(wb),
            entities=self._load_entities(wb),
            properties=self._load_properties(wb),
            logics=self._load_logics(wb),
            relationships=self._load_relationships(wb),
        )
        wb.close()
        return data

    def _load_domains(self, wb) -> list[DomainDef]:
        ws = wb[SHEET_DOMAIN]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        result = []
        for row in rows:
            if not row or not row[0]:
                continue
            result.append(
                DomainDef(
                    fqn=str(row[0]).strip(),
                    name_cn=str(row[1]).strip() if row[1] else "",
                    name_en=str(row[2]).strip() if len(row) > 2 and row[2] else "",
                    parent_fqn=str(row[3]).strip() if len(row) > 3 and row[3] else None,
                    description=str(row[4]).strip() if len(row) > 4 and row[4] else "",
                    source=str(row[5]).strip() if len(row) > 5 and row[5] else "manual",
                    status=str(row[6]).strip() if len(row) > 6 and row[6] else "active",
                )
            )
        return result

    def _load_entities(self, wb) -> list[EntityDef]:
        ws = wb[SHEET_ENTITY]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        result = []
        for row in rows:
            if not row or not row[0]:
                continue
            result.append(
                EntityDef(
                    fqn=str(row[0]).strip(),
                    entity_type=str(row[1]).strip() if row[1] else "",
                    name_cn=str(row[2]).strip() if len(row) > 2 and row[2] else "",
                    name_en=str(row[3]).strip() if len(row) > 3 and row[3] else "",
                    src_tables=self._parse_json(row[4]) if len(row) > 4 and row[4] else [],
                    domains=self._parse_json(row[5]) if len(row) > 5 and row[5] else [],
                    description=str(row[6]).strip() if len(row) > 6 and row[6] else "",
                    source=str(row[7]).strip() if len(row) > 7 and row[7] else "manual",
                    status=str(row[8]).strip() if len(row) > 8 and row[8] else "active",
                )
            )
        return result

    def _load_properties(self, wb) -> list[PropertyDef]:
        ws = wb[SHEET_PROPERTY]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        result = []
        for row in rows:
            if not row or not row[0]:
                continue
            result.append(
                PropertyDef(
                    fqn=str(row[0]).strip(),
                    entity_fqn=str(row[1]).strip() if row[1] else "",
                    data_type=str(row[2]).strip() if row[2] else "unknown",
                    is_pk=self._parse_bool(row[3]) if len(row) > 3 else False,
                    ref_property_fqn=str(row[4]).strip() if len(row) > 4 and row[4] else None,
                    description=str(row[5]).strip() if len(row) > 5 and row[5] else "",
                    name_cn=str(row[6]).strip() if len(row) > 6 and row[6] else "",
                    name_en=str(row[7]).strip() if len(row) > 7 and row[7] else "",
                    source=str(row[8]).strip() if len(row) > 8 and row[8] else "manual",
                    status=str(row[9]).strip() if len(row) > 9 and row[9] else "active",
                )
            )
        return result

    def _load_logics(self, wb) -> list[LogicDef]:
        if SHEET_LOGIC not in wb.sheetnames:
            return []
        ws = wb[SHEET_LOGIC]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        result = []
        for row in rows:
            if not row or not row[0]:
                continue
            result.append(
                LogicDef(
                    fqn=str(row[0]).strip(),
                    logic_type=str(row[1]).strip() if row[1] else "formula",
                    expression=str(row[2]).strip() if len(row) > 2 and row[2] else "",
                    name_cn=str(row[3]).strip() if len(row) > 3 and row[3] else "",
                    name_en=str(row[4]).strip() if len(row) > 4 and row[4] else "",
                    description=str(row[5]).strip() if len(row) > 5 and row[5] else "",
                    source=str(row[6]).strip() if len(row) > 6 and row[6] else "manual",
                    status=str(row[7]).strip() if len(row) > 7 and row[7] else "active",
                )
            )
        return result

    def _load_relationships(self, wb) -> list[RelationshipDef]:
        ws = wb[SHEET_RELATIONSHIP]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        result = []
        for row in rows:
            if not row or not row[0]:
                continue
            result.append(
                RelationshipDef(
                    src_fqn=str(row[0]).strip(),
                    dst_fqn=str(row[1]).strip() if row[1] else "",
                    rel_type=str(row[2]).strip() if len(row) > 2 and row[2] else "REFERENCES",
                    is_directed=self._parse_bool(row[3]) if len(row) > 3 else True,
                    source=str(row[4]).strip() if len(row) > 4 and row[4] else "introspect",
                    status=str(row[5]).strip() if len(row) > 5 and row[5] else "active",
                )
            )
        return result

    @staticmethod
    def _parse_bool(value) -> bool:
        if isinstance(value, bool):
            return value
        if isinstance(value, int):
            return bool(value)
        s = str(value).strip().lower()
        return s in ("true", "yes", "1", "是", "y")

    @staticmethod
    def _parse_json(value: str | None) -> Any:
        if not value:
            return {}
        if isinstance(value, dict):
            return value
        try:
            return json.loads(str(value))
        except (json.JSONDecodeError, TypeError):
            logger.warning("Failed to parse JSON properties: %s", value)
            return {}
